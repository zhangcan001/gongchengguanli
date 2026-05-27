from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


TARGET_FIELDS = (
    "building",
    "floor",
    "area",
    "discipline",
    "task_name",
    "unit",
    "total_quantity",
    "cumulative_quantity",
    "period_quantity",
    "weight",
    "planned_percent",
    "actual_percent",
    "planned_start_date",
    "planned_finish_date",
    "remark",
)


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "building": ("楼栋", "楼号", "单体", "楼座"),
    "floor": ("楼层", "层", "施工层"),
    "area": ("区域", "部位", "施工部位"),
    "discipline": ("专业", "工种", "分部", "类型"),
    "task_name": ("任务", "任务名称", "分部分项", "工程内容", "施工内容", "分项工程", "工作内容"),
    "unit": ("单位", "计量单位"),
    "total_quantity": ("总量", "工程量", "合同量", "总工程量"),
    "cumulative_quantity": ("累计完成", "累计完成量", "已完成量"),
    "period_quantity": ("本期完成", "本周完成", "今日完成", "本月完成"),
    "weight": ("权重", "weight", "占比", "统计权重", "权重系数", "系数"),
    "planned_percent": ("计划完成率", "计划进度", "计划百分比"),
    "actual_percent": ("实际完成率", "实际进度", "完成率", "形象进度", "完成进度"),
    "planned_start_date": ("计划开始", "计划开始日期", "开始日期"),
    "planned_finish_date": ("计划完成", "计划完成日期", "计划结束", "完成日期"),
    "remark": ("备注", "说明", "备注说明"),
}


FIELD_LABELS = {
    "building": "楼栋",
    "floor": "楼层",
    "area": "施工部位",
    "discipline": "专业",
    "task_name": "任务名称",
    "unit": "单位",
    "total_quantity": "总量",
    "cumulative_quantity": "累计完成量",
    "period_quantity": "本期完成量",
    "weight": "权重",
    "planned_percent": "计划完成率",
    "actual_percent": "实际完成率",
    "planned_start_date": "计划开始日期",
    "planned_finish_date": "计划完成日期",
    "remark": "备注",
}


SKIP_ROW_KEYWORDS = ("合计", "总计", "小计", "总 计", "小 计")
TITLE_OR_NOTE_KEYWORDS = (
    "说明",
    "备注：",
    "备注:",
    "注：",
    "注:",
    "填报",
    "编制",
    "审核",
    "项目名称",
    "工程名称",
    "单位：",
    "单位:",
)
HEADER_PARENT_KEYWORDS = (
    "楼栋",
    "楼层",
    "部位",
    "专业",
    "任务",
    "工程",
    "施工",
    "工程量",
    "完成",
    "累计",
    "本期",
    "计划",
    "实际",
    "进度",
    "权重",
    "日期",
)
DATE_PATTERN = re.compile(r"(20\d{2})\s*[-年/.]\s*(\d{1,2})\s*[-月/.]\s*(\d{1,2})\s*日?")
COMPACT_DATE_PATTERN = re.compile(r"(20\d{2})(\d{2})(\d{2})")
MONTH_DAY_PATTERN = re.compile(r"(?<!\d)(\d{1,2})\s*月\s*(\d{1,2})\s*日?")
NUMBER_PATTERN = re.compile(r"[-+]?\d+(?:\.\d+)?")
EMPTY_NUMBER_TOKENS = {"", "-", "--", "—", "–", "/", "\\", "无", "空"}
CODE_FIELD_KEYWORDS = ("编码", "编号", "code", "清单号", "序号")
NUMERIC_FIELDS = {"total_quantity", "cumulative_quantity", "period_quantity", "weight"}
PERCENT_FIELDS = {"planned_percent", "actual_percent"}
DATE_FIELDS = {"planned_start_date", "planned_finish_date"}


@dataclass
class FieldMappingDraft:
    source_field: str
    target_field: str
    confidence: float
    is_confirmed: bool


@dataclass
class ValidationIssueDraft:
    row_index: int
    field: str | None
    message: str
    severity: str


@dataclass
class PreviewRowDraft:
    row_index: int
    source: dict[str, Any]
    normalized: dict[str, Any]
    issues: list[ValidationIssueDraft]


@dataclass
class ImportStatsDraft:
    raw_row_count: int = 0
    skipped_row_count: int = 0
    importable_row_count: int = 0
    error_count: int = 0
    warning_count: int = 0


@dataclass
class HeaderDraft:
    source_field: str
    column_index: int


@dataclass
class ExcelAnalysisResult:
    sheet_name: str
    header_row_index: int
    data_start_row_index: int
    data_date: date
    field_mappings: list[FieldMappingDraft]
    preview_rows: list[PreviewRowDraft]
    warnings: list[ValidationIssueDraft]
    errors: list[ValidationIssueDraft]
    stats: ImportStatsDraft


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def is_empty_row(values: list[Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def should_skip_row(values: list[Any]) -> bool:
    if is_empty_row(values):
        return True
    text_values = [str(value).strip() for value in values if value is not None and str(value).strip()]
    text = "".join(text_values)
    if any(keyword in text for keyword in SKIP_ROW_KEYWORDS):
        return True
    if len(text_values) <= 2 and any(text.startswith(keyword) for keyword in TITLE_OR_NOTE_KEYWORDS):
        return True
    return False


def parse_date_value(value: Any, *, year_hint: int | None = None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 20000 <= float(value) <= 80000:
            try:
                return from_excel(value).date()
            except (TypeError, ValueError, OverflowError):
                return None
    if value is None:
        return None
    text = str(value).strip()
    match = DATE_PATTERN.search(text) or COMPACT_DATE_PATTERN.search(text)
    if match:
        year, month, day = (int(part) for part in match.groups())
        try:
            return date(year, month, day)
        except ValueError:
            return None

    month_day_match = MONTH_DAY_PATTERN.search(text)
    if not month_day_match:
        return None
    year = year_hint or date.today().year
    month, day = (int(part) for part in month_day_match.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_number(value: Any, *, percent: bool = False) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        has_percent_sign = False
    else:
        text = str(value).strip().replace(",", "")
        if text in EMPTY_NUMBER_TOKENS:
            return None
        has_percent_sign = "%" in text or "％" in text
        match = NUMBER_PATTERN.search(text)
        if not match:
            return None
        try:
            number = float(match.group(0))
        except ValueError:
            return None
    if percent and not has_percent_sign and 0 <= number <= 1:
        return round(number * 100, 6)
    return number


def has_meaningful_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() in EMPTY_NUMBER_TOKENS:
        return False
    return str(value).strip() != ""


def serialize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def match_field(source_field: str) -> tuple[str, float]:
    normalized = normalize_header(source_field)
    if any(keyword in normalized for keyword in CODE_FIELD_KEYWORDS):
        return "", 0
    priority_aliases: tuple[tuple[str, tuple[str, ...]], ...] = (
        ("planned_percent", ("计划完成率", "计划进度", "计划百分比")),
        ("actual_percent", ("实际完成率", "实际进度", "形象进度", "完成进度")),
        ("cumulative_quantity", ("累计完成量", "累计完成", "已完成量")),
        ("period_quantity", ("本期完成", "本周完成", "今日完成", "本月完成")),
        ("total_quantity", ("合同量", "总量", "总工程量", "工程量")),
        ("area", ("施工部位", "施工区域", "部位")),
        ("task_name", ("分部分项", "任务名称", "工作内容", "工程内容", "施工内容", "分项工程")),
    )
    for target_field, aliases in priority_aliases:
        for alias in aliases:
            normalized_alias = normalize_header(alias)
            if normalized == normalized_alias:
                return target_field, 0.98
            if normalized_alias and normalized_alias in normalized:
                return target_field, 0.88
    for target_field, aliases in FIELD_ALIASES.items():
        if normalized == target_field.lower():
            return target_field, 1
        for alias in aliases:
            normalized_alias = normalize_header(alias)
            if normalized == normalized_alias:
                return target_field, 0.98
            if normalized_alias and normalized_alias in normalized:
                return target_field, 0.86
    return "", 0


class ExcelAnalysisService:
    def analyze(self, file_path: Path, *, fallback_date: date | None = None) -> ExcelAnalysisResult:
        workbook = load_workbook(file_path, data_only=True, read_only=False)
        try:
            worksheet = self._select_sheet(workbook.worksheets)
            merged_lookup = self._build_merged_lookup(worksheet)
            header_row_index = self._detect_header_row(worksheet)
            headers = self._read_headers(worksheet, header_row_index, merged_lookup)
            data_start_row_index = self._detect_data_start_row(worksheet, header_row_index)
            mappings = self._build_mappings(headers)
            data_date = self._detect_data_date(file_path.name, worksheet, headers, data_start_row_index, fallback_date)
            preview_rows, warnings, errors, stats = self._build_preview(worksheet, headers, mappings, data_start_row_index, data_date.year)

            return ExcelAnalysisResult(
                sheet_name=worksheet.title,
                header_row_index=header_row_index,
                data_start_row_index=data_start_row_index,
                data_date=data_date,
                field_mappings=mappings,
                preview_rows=preview_rows,
                warnings=warnings,
                errors=errors,
                stats=stats,
            )
        finally:
            workbook.close()

    def validate(
        self,
        file_path: Path,
        *,
        sheet_name: str,
        header_row_index: int,
        data_start_row_index: int,
        mappings: list[FieldMappingDraft],
        year_hint: int | None = None,
    ) -> tuple[list[PreviewRowDraft], list[ValidationIssueDraft], list[ValidationIssueDraft], ImportStatsDraft]:
        workbook = load_workbook(file_path, data_only=True, read_only=False)
        try:
            worksheet = workbook[sheet_name]
            merged_lookup = self._build_merged_lookup(worksheet)
            headers = self._read_headers(worksheet, header_row_index, merged_lookup)
            return self._build_preview(worksheet, headers, mappings, data_start_row_index, year_hint)
        finally:
            workbook.close()

    def _build_merged_lookup(self, worksheet: Worksheet) -> dict[tuple[int, int], Any]:
        lookup: dict[tuple[int, int], Any] = {}
        for merged_range in worksheet.merged_cells.ranges:
            value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
            for row_index in range(merged_range.min_row, merged_range.max_row + 1):
                for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                    lookup[(row_index, column_index)] = value
        return lookup

    def _cell_value(self, worksheet: Worksheet, row_index: int, column_index: int, merged_lookup: dict[tuple[int, int], Any] | None = None) -> Any:
        value = worksheet.cell(row_index, column_index).value
        if value is not None:
            return value
        if merged_lookup:
            return merged_lookup.get((row_index, column_index))
        return None

    def _select_sheet(self, worksheets: list[Worksheet]) -> Worksheet:
        best_sheet = worksheets[0]
        best_score = -1
        for worksheet in worksheets:
            score = 0
            max_rows = min(worksheet.max_row or 0, 20)
            for row_index in range(1, max_rows + 1):
                row = [worksheet.cell(row_index, column_index).value for column_index in range(1, (worksheet.max_column or 0) + 1)]
                joined = " ".join(str(value) for value in row if value is not None)
                for aliases in FIELD_ALIASES.values():
                    if any(alias in joined for alias in aliases):
                        score += 1
            if score > best_score:
                best_score = score
                best_sheet = worksheet
        return best_sheet

    def _detect_header_row(self, worksheet: Worksheet) -> int:
        best_row = 1
        best_score = -1
        max_rows = min(worksheet.max_row or 0, 20)
        for row_index in range(1, max_rows + 1):
            row = [worksheet.cell(row_index, column_index).value for column_index in range(1, (worksheet.max_column or 0) + 1)]
            if is_empty_row(row):
                continue
            score = self._header_candidate_score(worksheet, row_index)
            if score == 0 and self._looks_like_title_or_note(row):
                continue
            if score > best_score or (score == best_score and row_index > best_row):
                best_score = score
                best_row = row_index
        return best_row

    def _header_candidate_score(self, worksheet: Worksheet, row_index: int) -> int:
        max_col = worksheet.max_column or 0
        leaf_score = 0
        combined_targets: set[str] = set()
        candidate_start = self._candidate_header_start(worksheet, row_index)
        for column_index in range(1, max_col + 1):
            if self._column_hidden(worksheet, column_index):
                continue
            leaf_value = worksheet.cell(row_index, column_index).value
            if leaf_value is not None:
                target_field, confidence = match_field(str(leaf_value))
                if target_field and confidence >= 0.8:
                    leaf_score += 1

            values: list[str] = []
            for header_row in range(candidate_start, row_index + 1):
                value = worksheet.cell(header_row, column_index).value
                if value is not None and str(value).strip():
                    values.append(str(value).strip())
            if not values:
                continue
            target_field, confidence = match_field(" ".join(values))
            if target_field and confidence >= 0.8:
                combined_targets.add(target_field)

        if leaf_score < 2:
            return 0
        return len(combined_targets) * 3 + leaf_score

    def _candidate_header_start(self, worksheet: Worksheet, row_index: int) -> int:
        start = row_index
        max_col = worksheet.max_column or 0
        for parent_row in range(row_index - 1, max(0, row_index - 3), -1):
            if parent_row < 1:
                break
            values = [worksheet.cell(parent_row, column_index).value for column_index in range(1, max_col + 1)]
            text = "".join(str(value).strip() for value in values if value is not None)
            non_empty_count = sum(1 for value in values if value is not None and str(value).strip())
            if non_empty_count == 0:
                break
            if non_empty_count <= 1 and self._looks_like_title_or_note(values):
                break
            if any(keyword in text for keyword in HEADER_PARENT_KEYWORDS):
                start = parent_row
                continue
            break
        return start

    def _read_headers(self, worksheet: Worksheet, header_row_index: int, merged_lookup: dict[tuple[int, int], Any] | None = None) -> list[HeaderDraft]:
        header_start = self._header_block_start(worksheet, header_row_index)
        headers: list[HeaderDraft] = []
        used_names: dict[str, int] = {}
        for column_index in range(1, (worksheet.max_column or 0) + 1):
            if self._column_hidden(worksheet, column_index):
                continue
            values: list[str] = []
            for row_index in range(header_start, header_row_index + 1):
                value = self._cell_value(worksheet, row_index, column_index, merged_lookup)
                if value is None:
                    continue
                text = str(value).strip()
                if text and text not in values:
                    values.append(text)
            if not values and self._column_is_empty(worksheet, column_index):
                continue
            header = " ".join(values).strip() if values else f"未命名列{column_index}"
            used_names[header] = used_names.get(header, 0) + 1
            if used_names[header] > 1:
                header = f"{header}(第{column_index}列)"
            headers.append(HeaderDraft(source_field=header, column_index=column_index))
        return headers

    def _detect_data_start_row(self, worksheet: Worksheet, header_row_index: int) -> int:
        for row_index in range(header_row_index + 1, (worksheet.max_row or header_row_index) + 1):
            row = [worksheet.cell(row_index, column_index).value for column_index in range(1, (worksheet.max_column or 0) + 1)]
            if not should_skip_row(row) and not self._looks_like_title_or_note(row):
                return row_index
        return header_row_index + 1

    def _build_mappings(self, headers: list[HeaderDraft]) -> list[FieldMappingDraft]:
        mappings: list[FieldMappingDraft] = []
        used_targets: set[str] = set()
        for header in headers:
            target_field, confidence = match_field(header.source_field)
            if target_field in used_targets:
                target_field = ""
                confidence = 0
            if target_field:
                used_targets.add(target_field)
            mappings.append(
                FieldMappingDraft(
                    source_field=header.source_field,
                    target_field=target_field,
                    confidence=confidence,
                    is_confirmed=False,
                )
            )
        return mappings

    def _detect_data_date(
        self,
        file_name: str,
        worksheet: Worksheet,
        headers: list[HeaderDraft],
        data_start_row_index: int,
        fallback_date: date | None,
    ) -> date:
        year_hint = fallback_date.year if fallback_date else None
        from_file_name = parse_date_value(file_name, year_hint=year_hint)
        if from_file_name:
            return from_file_name

        max_title_rows = min(data_start_row_index, 5)
        for row_index in range(1, max_title_rows + 1):
            row = [worksheet.cell(row_index, column_index).value for column_index in range(1, (worksheet.max_column or 0) + 1)]
            parsed = parse_date_value(" ".join(str(value) for value in row if value is not None), year_hint=year_hint)
            if parsed:
                return parsed

        date_columns = [
            header.column_index for header in headers if "日期" in header.source_field or "date" in header.source_field.lower()
        ]
        for row_index in range(data_start_row_index, min(data_start_row_index + 10, worksheet.max_row or data_start_row_index) + 1):
            for column_index in date_columns:
                parsed = parse_date_value(worksheet.cell(row_index, column_index).value, year_hint=year_hint)
                if parsed:
                    return parsed

        return fallback_date or date.today()

    def _build_preview(
        self,
        worksheet: Worksheet,
        headers: list[HeaderDraft],
        mappings: list[FieldMappingDraft],
        data_start_row_index: int,
        year_hint: int | None = None,
    ) -> tuple[list[PreviewRowDraft], list[ValidationIssueDraft], list[ValidationIssueDraft], ImportStatsDraft]:
        preview_rows: list[PreviewRowDraft] = []
        warnings: list[ValidationIssueDraft] = []
        errors: list[ValidationIssueDraft] = []
        mapped_by_source = {mapping.source_field: mapping.target_field for mapping in mappings if mapping.target_field}
        stats = ImportStatsDraft()
        merged_lookup = self._build_merged_lookup(worksheet)

        for row_index in range(data_start_row_index, (worksheet.max_row or data_start_row_index) + 1):
            stats.raw_row_count += 1
            values = [self._cell_value(worksheet, row_index, column_index, merged_lookup) for column_index in range(1, (worksheet.max_column or 0) + 1)]
            if should_skip_row(values):
                stats.skipped_row_count += 1
                continue

            source = {
                header.source_field: serialize_value(values[header.column_index - 1]) if header.column_index - 1 < len(values) else None
                for header in headers
            }
            normalized, parse_issues = self._normalize_row(row_index, source, mapped_by_source, year_hint=year_hint)
            if not self._row_has_importable_content(normalized) or self._looks_like_title_or_note(values):
                stats.skipped_row_count += 1
                continue
            row_issues = parse_issues + self._validate_row(row_index, normalized)
            preview = PreviewRowDraft(row_index=row_index, source=source, normalized=normalized, issues=row_issues)
            preview_rows.append(preview)
            for issue in row_issues:
                if issue.severity == "error":
                    errors.append(issue)
                else:
                    warnings.append(issue)

        stats.importable_row_count = len(preview_rows)
        stats.error_count = len(errors)
        stats.warning_count = len(warnings)
        return preview_rows, warnings, errors, stats

    def _normalize_row(
        self,
        row_index: int,
        source: dict[str, Any],
        mapped_by_source: dict[str, str],
        *,
        year_hint: int | None = None,
    ) -> tuple[dict[str, Any], list[ValidationIssueDraft]]:
        normalized = {field: None for field in TARGET_FIELDS}
        issues: list[ValidationIssueDraft] = []
        for source_field, value in source.items():
            target_field = mapped_by_source.get(source_field)
            if not target_field:
                continue
            if target_field in NUMERIC_FIELDS:
                parsed_number = parse_number(value)
                normalized[target_field] = parsed_number
                if parsed_number is None and has_meaningful_value(value):
                    issues.append(self._parse_issue(row_index, target_field, value, "数值"))
            elif target_field in PERCENT_FIELDS:
                parsed_percent = parse_number(value, percent=True)
                normalized[target_field] = parsed_percent
                if parsed_percent is None and has_meaningful_value(value):
                    issues.append(self._parse_issue(row_index, target_field, value, "百分比"))
            elif target_field in DATE_FIELDS:
                parsed = parse_date_value(value, year_hint=year_hint)
                normalized[target_field] = parsed.isoformat() if parsed else None
                if parsed is None and has_meaningful_value(value):
                    issues.append(self._parse_issue(row_index, target_field, value, "日期"))
            else:
                normalized[target_field] = serialize_value(value)
        return normalized, issues

    def _validate_row(self, row_index: int, normalized: dict[str, Any]) -> list[ValidationIssueDraft]:
        issues: list[ValidationIssueDraft] = []
        if normalized.get("task_name") in (None, ""):
            issues.append(ValidationIssueDraft(row_index, "task_name", f"第 {row_index} 行“任务名称”不能为空。", "error"))

        for field in ("actual_percent", "planned_percent"):
            value = normalized.get(field)
            if value is not None and (not isinstance(value, (int, float)) or value < 0 or value > 100):
                label = FIELD_LABELS[field]
                issues.append(ValidationIssueDraft(row_index, field, f"第 {row_index} 行“{label}”为 {value}，必须在 0 到 100 之间。", "error"))

        total_quantity = normalized.get("total_quantity")
        cumulative_quantity = normalized.get("cumulative_quantity")
        if total_quantity is not None and cumulative_quantity is not None and cumulative_quantity > total_quantity:
            issues.append(ValidationIssueDraft(row_index, "cumulative_quantity", f"第 {row_index} 行“累计完成量”为 {cumulative_quantity}，大于“总量” {total_quantity}。", "error"))

        period_quantity = normalized.get("period_quantity")
        if period_quantity is not None and period_quantity < 0:
            issues.append(ValidationIssueDraft(row_index, "period_quantity", f"第 {row_index} 行“本期完成量”为 {period_quantity}，不得小于 0。", "error"))

        for field in ("planned_start_date", "planned_finish_date"):
            value = normalized.get(field)
            if value:
                parsed = parse_date_value(value)
                if not parsed:
                    issues.append(ValidationIssueDraft(row_index, field, f"第 {row_index} 行“{FIELD_LABELS[field]}”填写为“{value}”，日期无法解析。", "error"))
                else:
                    normalized[field] = parsed.isoformat()

        if normalized.get("actual_percent") is None and normalized.get("planned_percent") is not None:
            issues.append(ValidationIssueDraft(row_index, "actual_percent", f"第 {row_index} 行缺少“实际完成率”，进度看板将标记为 no_calculable_progress。", "warning"))
        if normalized.get("planned_percent") is None and normalized.get("actual_percent") is not None:
            issues.append(ValidationIssueDraft(row_index, "planned_percent", f"第 {row_index} 行缺少“计划完成率”，无法判断进度滞后。", "warning"))

        return issues

    def _parse_issue(self, row_index: int, field: str, value: Any, value_type: str) -> ValidationIssueDraft:
        return ValidationIssueDraft(
            row_index,
            field,
            f"第 {row_index} 行“{FIELD_LABELS[field]}”填写为“{value}”，无法解析为{value_type}。",
            "error",
        )

    def _column_hidden(self, worksheet: Worksheet, column_index: int) -> bool:
        return bool(worksheet.column_dimensions[get_column_letter(column_index)].hidden)

    def _column_is_empty(self, worksheet: Worksheet, column_index: int) -> bool:
        max_row = min(worksheet.max_row or 0, 100)
        for row_index in range(1, max_row + 1):
            value = worksheet.cell(row_index, column_index).value
            if value is not None and str(value).strip():
                return False
        return True

    def _header_block_start(self, worksheet: Worksheet, header_row_index: int) -> int:
        start = header_row_index
        max_col = worksheet.max_column or 0
        for row_index in range(header_row_index - 1, max(0, header_row_index - 3), -1):
            if row_index < 1:
                break
            values = [worksheet.cell(row_index, column_index).value for column_index in range(1, max_col + 1)]
            text = "".join(str(value).strip() for value in values if value is not None)
            non_empty_count = sum(1 for value in values if value is not None and str(value).strip())
            if non_empty_count == 0:
                break
            if non_empty_count <= 1 and self._looks_like_title_or_note(values):
                break
            if any(keyword in text for keyword in HEADER_PARENT_KEYWORDS):
                start = row_index
                continue
            break
        return start

    def _looks_like_title_or_note(self, values: list[Any]) -> bool:
        text_values = [str(value).strip() for value in values if value is not None and str(value).strip()]
        if not text_values:
            return True
        joined = "".join(text_values)
        if len(text_values) <= 2 and any(keyword in joined for keyword in TITLE_OR_NOTE_KEYWORDS):
            return True
        if len(text_values) == 1 and any(keyword in text_values[0] for keyword in ("进度表", "进度统计", "统计表", "汇总表")):
            return True
        return False

    def _row_has_importable_content(self, normalized: dict[str, Any]) -> bool:
        meaningful_fields = [
            "building",
            "floor",
            "area",
            "discipline",
            "task_name",
            "total_quantity",
            "cumulative_quantity",
            "period_quantity",
            "planned_percent",
            "actual_percent",
        ]
        return any(has_meaningful_value(normalized.get(field)) for field in meaningful_fields)


def mapping_to_dict(mapping: FieldMappingDraft) -> dict[str, Any]:
    return {
        "source_field": mapping.source_field,
        "target_field": mapping.target_field,
        "confidence": mapping.confidence,
        "is_confirmed": mapping.is_confirmed,
    }


def issue_to_dict(issue: ValidationIssueDraft) -> dict[str, Any]:
    return {
        "row_index": issue.row_index,
        "field": issue.field,
        "message": issue.message,
        "severity": issue.severity,
    }


def stats_to_dict(stats: ImportStatsDraft) -> dict[str, int]:
    return {
        "raw_row_count": stats.raw_row_count,
        "skipped_row_count": stats.skipped_row_count,
        "importable_row_count": stats.importable_row_count,
        "error_count": stats.error_count,
        "warning_count": stats.warning_count,
    }


def preview_to_dict(row: PreviewRowDraft) -> dict[str, Any]:
    return {
        "row_index": row.row_index,
        "source": row.source,
        "normalized": row.normalized,
        "issues": [issue_to_dict(issue) for issue in row.issues],
    }

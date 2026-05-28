from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from app.excel_analysis import ExcelAnalysisService, match_field, parse_date_value, parse_number
from tests.test_smart_inbox import create_project


def make_progress_workbook(*, invalid: bool = False, data_date: str = "2026-05-26") -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "本周进度"
    worksheet.append([f"项目进度统计 {data_date}"])
    worksheet.append(["楼栋", "楼层", "施工部位", "专业", "任务名称", "单位", "总工程量", "累计完成量", "本周完成", "计划完成率", "实际完成率", "计划开始", "计划完成", "备注"])
    if invalid:
        worksheet.append(["3#楼", "12层", "砌体", "土建", "", "m2", 100, 120, -1, 80, 135, "bad-date", "2026-06-30", "异常行"])
    else:
        worksheet.append(["3#楼", "12层", "砌体", "土建", "砌体施工", "m2", 100, 80, 20, 75, 80, "2026-05-01", "2026-06-30", "正常"])
        worksheet.append(["合计", None, None, None, None, None, 100, 80, 20, None, None, None, None, None])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def upload_progress_excel(client, project_id: int, *, content: bytes | None = None, file_name: str = "进度表_2026-05-26.xlsx"):
    return client.post(
        "/api/smart-inbox/upload",
        data={"project_id": str(project_id)},
        files={
            "file": (
                file_name,
                content if content is not None else make_progress_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_excel_analysis_reads_sheet_header_mapping_and_preview(tmp_path: Path):
    excel_path = tmp_path / "进度表_2026-05-26.xlsx"
    excel_path.write_bytes(make_progress_workbook())

    result = ExcelAnalysisService().analyze(excel_path)

    assert result.sheet_name == "本周进度"
    assert result.header_row_index == 2
    assert result.data_start_row_index == 3
    assert result.data_date.isoformat() == "2026-05-26"
    assert any(mapping.target_field == "task_name" for mapping in result.field_mappings)
    assert len(result.preview_rows) == 1
    assert result.preview_rows[0].normalized["task_name"] == "砌体施工"
    assert result.preview_rows[0].normalized["actual_percent"] == 80


def test_field_matching_uses_aliases_without_hardcoding_columns():
    assert match_field("楼栋") == ("building", 0.98)
    assert match_field("形象进度") == ("actual_percent", 0.98)
    assert match_field("工程内容") == ("task_name", 0.98)
    assert match_field("计划完成率") == ("planned_percent", 0.98)
    assert match_field("施工部位") == ("area", 0.98)
    assert match_field("分部分项") == ("task_name", 0.98)
    assert match_field("任务编码") == ("", 0)
    assert match_field("清单编码") == ("", 0)
    assert match_field("未知字段") == ("", 0)


def test_percent_number_and_date_parsing_real_formats():
    assert parse_number("80%", percent=True) == 80
    assert parse_number("80.00%", percent=True) == 80
    assert parse_number(0.8, percent=True) == 80
    assert parse_number("80", percent=True) == 80
    assert parse_number("约1,200.5㎡") == 1200.5
    assert parse_number("-") is None
    assert parse_number("未填") is None
    assert parse_date_value("2026-05-26").isoformat() == "2026-05-26"
    assert parse_date_value("2026/5/26").isoformat() == "2026-05-26"
    assert parse_date_value("5月26日", year_hint=2026).isoformat() == "2026-05-26"
    assert parse_date_value(to_excel(datetime(2026, 5, 26))).isoformat() == "2026-05-26"


def test_excel_validation_reports_errors(tmp_path: Path):
    excel_path = tmp_path / "进度表_2026-05-26.xlsx"
    excel_path.write_bytes(make_progress_workbook(invalid=True))

    result = ExcelAnalysisService().analyze(excel_path)
    error_fields = {issue.field for issue in result.errors}

    assert "task_name" in error_fields
    assert "actual_percent" in error_fields
    assert "cumulative_quantity" in error_fields
    assert "period_quantity" in error_fields
    assert "planned_start_date" in error_fields
    assert any("第 3 行" in issue.message and "实际完成率" in issue.message for issue in result.errors)


def make_multi_sheet_multi_header_workbook() -> bytes:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "封面"
    cover.append(["项目进度说明"])
    cover.append(["本 sheet 不含进度明细"])

    worksheet = workbook.create_sheet("真实进度")
    worksheet.append(["某项目进度统计表"])
    worksheet.append(["数据日期：5月26日"])
    worksheet.append(["任务编码", "楼栋", "楼层", "施工部位", "分部分项", "单位", "工程量", "工程量", "工程量", "进度", "进度", "备注"])
    worksheet.append(["", "", "", "", "工作内容", "", "合同量", "累计完成量", "本周完成", "计划完成率", "形象进度", ""])
    worksheet.column_dimensions["A"].hidden = True
    worksheet.append(["QD-001", "1#楼", "3层", "东单元", "砌体施工", "m2", "100㎡", "80㎡", "20㎡", "80%", 0.8, "正常"])
    worksheet.append(["QD-002", "1#楼", "4层", "西单元", "抹灰施工", "m2", "200㎡", "120㎡", "40㎡", 70, "60.00%", "轻微滞后"])
    worksheet.append(["合计", None, None, None, None, None, 300, 200, 60, None, None, None])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_merged_header_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "合并表头"
    worksheet.append(["进度统计 2026/5/26"])
    worksheet.append(["楼栋", "楼层", "任务名称", "工程量", None, None, "进度", None, "计划开始", "计划完成"])
    worksheet.merge_cells("D2:F2")
    worksheet.merge_cells("G2:H2")
    worksheet.append(["楼栋", "楼层", "工作内容", "合同量", "累计完成量", "今日完成", "计划完成率", "实际完成率", "计划开始", "计划完成"])
    worksheet.append(["2#楼", "1层", "钢筋绑扎", "100t", "60t", "10t", "75%", "70%", to_excel(datetime(2026, 5, 1)), "6月30日"])
    worksheet.append(["小计", None, None, 100, 60, 10, None, None, None, None])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_realistic_plan_actual_workbook() -> bytes:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "填报说明"
    cover.append(["进度表上传说明"])
    cover.append(["本页包含计划完成率、实际完成率、累计完成量等填写说明，不是进度明细。"])
    cover.append(["请以明细 sheet 为准。"])

    worksheet = workbook.create_sheet("月度进度")
    worksheet.append(["智能工程项目月度进度表"])
    worksheet.append(["施工单位：测试施工单位"])
    worksheet.append(["楼栋", "楼层", "任务", "计划", "实际", None, "计划日期", None, "备注"])
    worksheet.merge_cells("E3:F3")
    worksheet.merge_cells("G3:H3")
    worksheet.append(["楼栋", "楼层", "工作内容", "完成率", "完成率", "累计完成量", "计划开始", "计划完成", "备注"])
    worksheet.append(["", "", "", "%", "%", "m2", "日期", "日期", ""])
    worksheet.append(["5#楼", "2层", "砌体施工", 0.75, "80.00%", "80㎡", "5月1日", to_excel(datetime(2026, 6, 30)), "正常推进"])
    worksheet.append(["总计", None, None, None, None, "80㎡", None, None, None])
    worksheet["D6"].number_format = "0.00%"

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_invalid_values_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "异常数据"
    worksheet.append(["进度表 2026-05-26"])
    worksheet.append(["楼栋", "任务名称", "合同量", "累计完成量", "本期完成", "计划完成率", "实际完成率", "计划开始"])
    worksheet.append(["3#楼", "模板安装", "100m2", "abc", "—", "80%", "完成八成", "5月26日"])
    worksheet.append(["3#楼", "混凝土浇筑", "100m3", "120m3", -1, "110%", "90%", "bad-date"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_missing_plan_actual_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "缺少字段"
    worksheet.append(["进度统计 2026-05-26"])
    worksheet.append(["楼栋", "分部分项", "实际完成率"])
    worksheet.append(["1#楼", "砌体施工", "80%"])
    worksheet.append(["2#楼", "抹灰施工", None])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_multi_sheet_multi_header_hidden_columns_and_skip_rows(tmp_path: Path):
    excel_path = tmp_path / "真实样表.xlsx"
    excel_path.write_bytes(make_multi_sheet_multi_header_workbook())

    result = ExcelAnalysisService().analyze(excel_path, fallback_date=datetime(2026, 5, 27).date())
    targets = {mapping.source_field: mapping.target_field for mapping in result.field_mappings}

    assert result.sheet_name == "真实进度"
    assert result.header_row_index == 4
    assert result.data_start_row_index == 5
    assert result.data_date.isoformat() == "2026-05-26"
    assert "任务编码" not in targets
    assert targets["施工部位"] == "area"
    assert targets["分部分项 工作内容"] == "task_name"
    assert targets["工程量 合同量"] == "total_quantity"
    assert targets["工程量 累计完成量"] == "cumulative_quantity"
    assert targets["工程量 本周完成"] == "period_quantity"
    assert targets["进度 计划完成率"] == "planned_percent"
    assert targets["进度 形象进度"] == "actual_percent"
    assert len(result.preview_rows) == 2
    assert result.preview_rows[0].normalized["actual_percent"] == 80
    assert result.stats.raw_row_count == 3
    assert result.stats.skipped_row_count == 1
    assert result.stats.importable_row_count == 2


def test_merged_header_and_excel_serial_dates_are_supported(tmp_path: Path):
    excel_path = tmp_path / "合并表头_2026-05-26.xlsx"
    excel_path.write_bytes(make_merged_header_workbook())

    result = ExcelAnalysisService().analyze(excel_path)
    row = result.preview_rows[0].normalized

    assert result.header_row_index == 3
    assert len(result.preview_rows) == 1
    assert row["total_quantity"] == 100
    assert row["cumulative_quantity"] == 60
    assert row["period_quantity"] == 10
    assert row["planned_start_date"] == "2026-05-01"
    assert row["planned_finish_date"] == "2026-06-30"


def test_realistic_cover_unit_row_and_plan_actual_headers(tmp_path: Path):
    excel_path = tmp_path / "现场月度进度_2026-05-26.xlsx"
    excel_path.write_bytes(make_realistic_plan_actual_workbook())

    result = ExcelAnalysisService().analyze(excel_path)
    targets = {mapping.source_field: mapping.target_field for mapping in result.field_mappings}
    row = result.preview_rows[0].normalized

    assert result.sheet_name == "月度进度"
    assert result.header_row_index == 4
    assert result.data_start_row_index == 6
    assert result.data_date.isoformat() == "2026-05-26"
    assert targets["计划 完成率"] == "planned_percent"
    assert targets["实际 完成率"] == "actual_percent"
    assert row["planned_percent"] == 75
    assert row["actual_percent"] == 80
    assert row["planned_start_date"] == "2026-05-01"
    assert row["planned_finish_date"] == "2026-06-30"
    assert result.stats.raw_row_count == 2
    assert result.stats.skipped_row_count == 1
    assert result.stats.importable_row_count == 1


def test_invalid_percent_date_and_number_have_precise_messages(tmp_path: Path):
    excel_path = tmp_path / "异常数据_2026-05-26.xlsx"
    excel_path.write_bytes(make_invalid_values_workbook())

    result = ExcelAnalysisService().analyze(excel_path)
    messages = [issue.message for issue in result.errors]
    fields = {issue.field for issue in result.errors}

    assert "cumulative_quantity" in fields
    assert "actual_percent" in fields
    assert "planned_start_date" in fields
    assert "planned_percent" in fields
    assert "period_quantity" in fields
    assert any("第 3 行" in message and "累计完成量" in message and "abc" in message for message in messages)
    assert any("第 4 行" in message and "计划开始日期" in message and "bad-date" in message for message in messages)


def test_missing_plan_or_actual_warns_without_blocking_import(tmp_path: Path):
    excel_path = tmp_path / "缺少字段_2026-05-26.xlsx"
    excel_path.write_bytes(make_missing_plan_actual_workbook())

    result = ExcelAnalysisService().analyze(excel_path)
    warning_fields = {issue.field for issue in result.warnings}

    assert result.errors == []
    assert "planned_percent" in warning_fields
    assert result.preview_rows[1].normalized["actual_percent"] is None


def test_progress_import_analyze_creates_draft_batch(client):
    project = create_project(client)
    upload_response = upload_progress_excel(client, project["id"])
    assert upload_response.status_code == 200
    inbox_id = upload_response.json()["inbox_id"]

    analyze_response = client.post(
        "/api/progress/import/analyze",
        json={"project_id": project["id"], "inbox_id": inbox_id},
    )

    assert analyze_response.status_code == 200
    payload = analyze_response.json()
    assert payload["batch_id"] > 0
    assert payload["detected_sheet"] == "本周进度"
    assert payload["data_date"] == "2026-05-26"
    assert payload["field_mappings"]
    assert payload["preview_rows"]
    assert payload["errors"] == []
    assert payload["import_stats"]["raw_row_count"] == 2
    assert payload["import_stats"]["skipped_row_count"] == 1
    assert payload["import_stats"]["importable_row_count"] == 1

    detail_response = client.get(f"/api/progress/import-batches/{payload['batch_id']}")
    assert detail_response.status_code == 200
    detail = detail_response.json()
    assert detail["status"] == "draft"
    assert detail["preview_rows"][0]["normalized"]["task_name"] == "砌体施工"


def test_progress_import_validate_updates_mapping(client):
    project = create_project(client)
    upload_response = upload_progress_excel(client, project["id"])
    inbox_id = upload_response.json()["inbox_id"]
    batch = client.post("/api/progress/import/analyze", json={"project_id": project["id"], "inbox_id": inbox_id}).json()
    mappings = batch["field_mappings"]
    for mapping in mappings:
        if mapping["source_field"] == "备注":
            mapping["target_field"] = ""
        mapping["is_confirmed"] = True

    response = client.post(f"/api/progress/import/{batch['batch_id']}/validate", json={"field_mappings": mappings})

    assert response.status_code == 200
    detail = response.json()
    assert detail["status"] == "validated"
    assert detail["validation_errors"] == []
    remark_mapping = next(mapping for mapping in detail["field_mappings"] if mapping["source_field"] == "备注")
    assert remark_mapping["target_field"] == ""


def test_confirmed_field_mapping_is_reused_on_later_analysis(client):
    project = create_project(client)
    first_upload = upload_progress_excel(client, project["id"])
    first_batch = client.post(
        "/api/progress/import/analyze",
        json={"project_id": project["id"], "inbox_id": first_upload.json()["inbox_id"]},
    ).json()
    mappings = first_batch["field_mappings"]
    for mapping in mappings:
        if mapping["source_field"] == "备注":
            mapping["target_field"] = ""
        mapping["is_confirmed"] = True

    validate_response = client.post(
        f"/api/progress/import/{first_batch['batch_id']}/validate",
        json={"field_mappings": mappings},
    )
    assert validate_response.status_code == 200

    second_upload = upload_progress_excel(client, project["id"])
    second_batch = client.post(
        "/api/progress/import/analyze",
        json={"project_id": project["id"], "inbox_id": second_upload.json()["inbox_id"]},
    ).json()

    remark_mapping = next(mapping for mapping in second_batch["field_mappings"] if mapping["source_field"] == "备注")
    assert remark_mapping["target_field"] == ""
    assert remark_mapping["is_confirmed"] is True


def test_progress_publish_writes_records_and_diary_material(client):
    project = create_project(client)
    upload_response = upload_progress_excel(client, project["id"])
    batch = client.post(
        "/api/progress/import/analyze",
        json={"project_id": project["id"], "inbox_id": upload_response.json()["inbox_id"]},
    ).json()

    publish_response = client.post(f"/api/progress/import/{batch['batch_id']}/publish", json={"replace_existing": False})

    assert publish_response.status_code == 200
    publish_payload = publish_response.json()
    assert publish_payload["status"] == "published"
    assert publish_payload["published_records"] == 1

    detail = client.get(f"/api/progress/import-batches/{batch['batch_id']}").json()
    assert detail["status"] == "published"


def test_publish_excludes_total_rows_from_progress_records(client):
    project = create_project(client)
    upload_response = upload_progress_excel(client, project["id"], content=make_multi_sheet_multi_header_workbook())
    batch = client.post(
        "/api/progress/import/analyze",
        json={"project_id": project["id"], "inbox_id": upload_response.json()["inbox_id"]},
    ).json()

    publish_response = client.post(f"/api/progress/import/{batch['batch_id']}/publish", json={"replace_existing": False})

    assert publish_response.status_code == 200
    assert publish_response.json()["published_records"] == 2


def test_sample_progress_workbooks_are_analyzable():
    sample_dir = Path(__file__).resolve().parents[2] / "resources" / "sample_data"
    samples = {
        "progress_standard_2026-05-26.xlsx": {"errors": 0, "rows": 1},
        "progress_multi_header_with_total_2026-05-26.xlsx": {"errors": 0, "rows": 2},
        "progress_invalid_date_percent_2026-05-26.xlsx": {"errors_min": 1, "rows": 2},
        "progress_missing_plan_actual_2026-05-26.xlsx": {"errors": 0, "rows": 2},
        "progress_realistic_plan_actual_2026-05-26.xlsx": {"errors": 0, "rows": 1},
    }

    for file_name, expectation in samples.items():
        result = ExcelAnalysisService().analyze(sample_dir / file_name, fallback_date=datetime(2026, 5, 27).date())
        assert result.stats.importable_row_count == expectation["rows"]
        if "errors" in expectation:
            assert len(result.errors) == expectation["errors"]
        else:
            assert len(result.errors) >= expectation["errors_min"]


def test_progress_publish_requires_replace_for_same_project_and_date(client):
    project = create_project(client)
    first_upload = upload_progress_excel(client, project["id"], file_name="进度表_2026-05-26.xlsx")
    first_batch = client.post(
        "/api/progress/import/analyze",
        json={"project_id": project["id"], "inbox_id": first_upload.json()["inbox_id"]},
    ).json()
    first_publish = client.post(f"/api/progress/import/{first_batch['batch_id']}/publish", json={"replace_existing": False})
    assert first_publish.status_code == 200

    second_upload = upload_progress_excel(client, project["id"], file_name="第二次进度表_2026-05-26.xlsx")
    second_analyze = client.post(
        "/api/progress/import/analyze",
        json={"project_id": project["id"], "inbox_id": second_upload.json()["inbox_id"]},
    )
    assert second_analyze.status_code == 200
    second_batch = second_analyze.json()
    assert second_batch["replacement_required"] is True

    blocked_publish = client.post(f"/api/progress/import/{second_batch['batch_id']}/publish", json={"replace_existing": False})
    assert blocked_publish.status_code == 409
    assert blocked_publish.json()["detail"]["code"] == "IMPORT_BATCH_REPLACEMENT_REQUIRED"

    replace_publish = client.post(f"/api/progress/import/{second_batch['batch_id']}/publish", json={"replace_existing": True})
    assert replace_publish.status_code == 200
    assert replace_publish.json()["replaced_existing"] is True
